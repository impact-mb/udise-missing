# app.py
# Streamlit app:
# 1. Upload Excel (all columns as string)
# 2. Ask for output folder
# 3. Keep only PROGRAMSUBTYPENAME == "ADOLOSCENT"
# 4. From those, keep rows with missing School UDISE
# 5. Optional: filter to ProgramLaunchName from a second file
# 6. Create data-quality flags (UDISE, phone, school, DOB, caste, parent consent)
# 7. Add footer rows (blank + filter description with timestamp + checks)
# 8. Export:
#    - One Excel with all issue records (ALL_CPRF_issues.xlsx)
#    - One Excel per ProgramLaunchName
# 9. Color-code cells with issues:
#    - >2 issues in a row: red (#FF0D0D)
#    - 1–2 issues: yellow (#FFFF66)
# 10. Show summary table + allow ZIP download of all files
# 11. Maintain a persistent run counter in run_counter.txt

import re
import io
import zipfile
from pathlib import Path
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

st.set_page_config(page_title="CPRF Data Quality Checker", layout="wide")

# -----------------------------
# Run counter (persistent in file)
# -----------------------------
COUNTER_FILE = Path(__file__).with_name("run_counter.txt")

def get_run_counter() -> int:
    """Read current run counter from file. Return 0 if file missing/invalid."""
    try:
        text = COUNTER_FILE.read_text(encoding="utf-8").strip()
        return int(text) if text != "" else 0
    except Exception:
        return 0

def increment_run_counter() -> int:
    """Increment run counter, save to file, and return new value."""
    current = get_run_counter()
    new_value = current + 1
    COUNTER_FILE.write_text(str(new_value), encoding="utf-8")
    return new_value

# Read current counter (for display)
initial_run_count = get_run_counter()

# -----------------------------
# UI header
# -----------------------------
st.title("CPRF Data Quality Checker – Adolescent Records")

# Show counter in sidebar
st.sidebar.header("Usage")
st.sidebar.metric("Total runs (all time)", initial_run_count)

st.markdown(
    """
1. Upload the raw CPRF Excel file.  
2. (Optional) Upload another Excel file that contains the list of **ProgramLaunchName** you want to filter on.  
3. Enter the folder path where you want output files to be saved.  
4. The app will:
   - Keep only rows where **PROGRAMSUBTYPENAME = "ADOLOSCENT"**  
   - From those, keep rows with **missing School UDISE**  
   - Optionally filter to selected **ProgramLaunchName** values  
   - Add data-quality flags:
     - **UDISE_Missing(Yes/No)** – `Missing` / `Available`  
     - **Check Child School Name** – child school name `Missing` / `Available`  
     - **Check Date Of Birth** – `Flag – DOB is 1 Jan` when applicable  
     - **Check Phone Number** – `Missing/Invalid` / `Valid`  
     - **CASTE_Interpretation** – friendly text for `DONT KNOW` / `DONT WISH`  
     - **Check Caste** – marks caste not known / not disclosed  
     - **Check Parent Consent** – `Missing` / `Available`  
   - Add a footer row with timestamp and list of checks (after one blank row)  
   - Export:
     - One combined Excel with all issue records  
     - Separate Excels by **ProgramLaunchName**  
   - Colour cells with issues (yellow/red based on number of issues in that row)  
   - Show a summary table and let you download all files as a **ZIP**

👉 To download the files, please go to the end of the page and click on **Download ZIP (ALL_CPRF_issues + ProgramLaunchName files)**.
"""
)

# -----------------------------
# Helper functions
# -----------------------------
def safe_filename(name: str) -> str:
    """Turn ProgramLaunchName into a safe filename."""
    if not isinstance(name, str):
        name = str(name)
    name = name.strip()
    # Keep alphanumeric, space, dash, underscore; replace others with "_"
    name = re.sub(r"[^\w\-\s]", "_", name)
    # Replace spaces with underscore
    name = re.sub(r"\s+", "_", name)
    if not name:
        name = "UnknownProgramLaunch"
    return name[:150]  # avoid extremely long names


def load_excel_as_string(file) -> pd.DataFrame:
    """Read Excel with all columns as string dtype (where possible)."""
    try:
        df = pd.read_excel(file, dtype="string", engine="openpyxl")
    except TypeError:
        # Fallback for older pandas that don't support dtype="string"
        df = pd.read_excel(file, engine="openpyxl")
        df = df.astype("string")
    return df


def get_programlaunch_list_from_file(file) -> list[str]:
    """
    Read an Excel file and extract a list of ProgramLaunchName values.

    Priority:
    1. Column whose name (after strip+lower) == 'programlaunchname'
    2. Otherwise, first column.
    """
    df_pl = load_excel_as_string(file)
    if df_pl.empty or len(df_pl.columns) == 0:
        return []

    # Try to find a column explicitly named ProgramLaunchName (case-insensitive)
    col_candidates = [
        c for c in df_pl.columns if c.strip().lower() == "programlaunchname"
    ]
    if col_candidates:
        col = col_candidates[0]
    else:
        # Fallback: use the first column
        col = df_pl.columns[0]

    values = (
        df_pl[col]
        .dropna()
        .astype(str)
        .str.strip()
    )
    values = [v for v in values if v != ""]
    return list(dict.fromkeys(values))  # unique while preserving order


def apply_issue_coloring(ws, data_df: pd.DataFrame):
    """
    Apply cell background colors based on issue counts per row.

    Logic:
    - Define issues using flag columns.
    - Map issues to original data columns:
        UDISE_Missing(Yes/No) -> "School UDISE"
        Check Child School Name -> "Child School Name"
        Check Date Of Birth -> "DATE OF BIRTH"
        Check Phone Number -> "CONTACTNUMBER"
        Check Caste -> "CASTE"
        Check Parent Consent -> "Parent Consent"
    - If >2 issues -> red (#FF0D0D)
    - If 1–2 issues -> yellow (#FFFF66)
    - Header is row 1 in Excel, data rows start at row 2.
    """

    # Define colors
    red_fill = PatternFill(start_color="FF0D0D", end_color="FF0D0D", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF66", end_color="FFFF66", fill_type="solid")

    cols = list(data_df.columns)
    col_idx_map = {col: i + 1 for i, col in enumerate(cols)}  # Excel is 1-based

    # Build masks based on existing flag columns
    masks = {}

    # UDISE
    if "UDISE_Missing(Yes/No)" in data_df.columns and "School UDISE" in data_df.columns:
        masks["School UDISE"] = data_df["UDISE_Missing(Yes/No)"].eq("Missing")
    # Child School Name
    if "Check Child School Name" in data_df.columns and "Child School Name" in data_df.columns:
        masks["Child School Name"] = data_df["Check Child School Name"].eq("Missing")
    # DOB
    if "Check Date Of Birth" in data_df.columns and "DATE OF BIRTH" in data_df.columns:
        masks["DATE OF BIRTH"] = data_df["Check Date Of Birth"].eq("Flag – DOB is 1 Jan")
    # Phone
    if "Check Phone Number" in data_df.columns and "CONTACTNUMBER" in data_df.columns:
        masks["CONTACTNUMBER"] = data_df["Check Phone Number"].eq("Missing/Invalid")
    # Caste
    if "Check Caste" in data_df.columns and "CASTE" in data_df.columns:
        masks["CASTE"] = data_df["Check Caste"].ne("")
    # Parent Consent
    if "Check Parent Consent" in data_df.columns and "Parent Consent" in data_df.columns:
        masks["Parent Consent"] = data_df["Check Parent Consent"].eq("Missing")

    n_rows = len(data_df)

    for i in range(n_rows):
        issue_cols = []
        for col_name, mask in masks.items():
            try:
                if bool(mask.iloc[i]):
                    issue_cols.append(col_name)
            except IndexError:
                continue

        issue_count = len(issue_cols)
        if issue_count == 0:
            continue

        # Choose color
        fill = red_fill if issue_count > 2 else yellow_fill

        excel_row = 2 + i  # row 2 is first data row
        for col_name in issue_cols:
            if col_name not in col_idx_map:
                continue
            col_index = col_idx_map[col_name]
            ws.cell(row=excel_row, column=col_index).fill = fill


# -----------------------------
# UI: Upload & output folder
# -----------------------------
uploaded_file = st.file_uploader(
    "Upload main CPRF Excel file (with delivery data)",
    type=["xlsx", "xls"],
    accept_multiple_files=False,
)

filter_file = st.file_uploader(
    "Optional: Upload Excel file that contains ProgramLaunchName list to filter on",
    type=["xlsx", "xls"],
    accept_multiple_files=False,
)

output_folder_input = st.text_input(
    "Output folder path (will be created if it doesn't exist)",
    value=str(Path.cwd() / "cprf_dq_output"),
)

run_button = st.button("Run and Save Files")

# -----------------------------
# Main logic
# -----------------------------
if run_button:
    if not uploaded_file:
        st.error("Please upload the main Excel file first.")
    else:
        # Load main data
        with st.spinner("Reading main Excel file..."):
            df = load_excel_as_string(uploaded_file)

        st.success(f"Main file loaded with {len(df):,} rows and {len(df.columns)} columns.")

        # Expected columns (as per your list)
        expected_columns = [
            "COUNTRYNAME",
            "REGIONNAME",
            "STATENAME",
            "DISTRICTNAME",
            "Community/School",
            "School Type",
            "School UDISE",
            "PROGRAMTYPENAME",
            "PROGRAMSUBTYPENAME",
            "Day Of Session",
            "Group Registration Date",
            "Session Timing",
            "YM NAME",
            "TMO NAME",
            "ProgramLaunchName",
            "FUNDERNAME",
            "ProjectName",
            "ProjectType",
            "GROUPID",
            "Group Status",
            "Child School Name",
            "CHILDID",
            "Intervention Year",
            "CHILDREGNO",
            "DATE OF JOINING",
            "FNAME",
            "MNAME",
            "LNAME",
            "GENDER",
            "ISDOBKNOWN",
            "DATE OF BIRTH",
            "AGE",
            "CHILDGOSCHOOL",
            "SCHOOLTYPENAME",
            "Class Of the Child Attending school",
            "CHILDDROPEDSCHOOL",
            "CLASSCHILDDROPEDSCHOOL",
            "REASONFORDROUPOUT",
            "CHILDDISABILITY",
            "DISABILITYNAME",
            "OTHERS",
            "WASPARTOFMBPROGRAM",
            "PREVIOUSCHILDREGNO",
            "REMARKS",
            "STATUS",
            "GUARDIAN",
            "P_Poverty Line(APL/BPL)",
            "CONTACTTYPE",
            "CONTACTNUMBER",
            "P_Do you Have Document?",
            "DOCUMENTTYPE",
            "DOCUMENTNO",
            "P_FName",
            "P_Age",
            "RELATION",
            "RELIGIONNAME",
            "CASTE",
            "TRIBE",
            "Previous year grade",
            "School Academic Cycle",
            "School HM/Teacher Contact",
            "Child Level",
            "Parent Consent",
        ]

        missing_cols = [c for c in expected_columns if c not in df.columns]
        if missing_cols:
            st.warning(
                "The following expected columns are missing from the uploaded main file:\n\n"
                + ", ".join(missing_cols)
            )

        # Core required columns
        required_cols = ["PROGRAMSUBTYPENAME", "School UDISE", "ProgramLaunchName"]
        if any(col not in df.columns for col in required_cols):
            st.error(
                "Required columns 'PROGRAMSUBTYPENAME', 'School UDISE', or 'ProgramLaunchName' "
                "are missing. Please check your main file."
            )
        else:
            # Clean key columns
            df["PROGRAMSUBTYPENAME"] = df["PROGRAMSUBTYPENAME"].str.strip()
            df["ProgramLaunchName"] = df["ProgramLaunchName"].astype("string").str.strip()

            # Step 1: adolescent filter
            adol = df[df["PROGRAMSUBTYPENAME"].str.upper() == "ADOLOSCENT"].copy()
            st.write(f"Rows with PROGRAMSUBTYPENAME = 'ADOLOSCENT': **{len(adol):,}**")

            # Step 2: missing School UDISE: blank or NA
            udise_col = adol["School UDISE"]
            missing_mask = udise_col.isna() | udise_col.str.strip().eq("")
            missing_df = adol[missing_mask].copy()

            st.write(
                f"Rows with missing 'School UDISE' among adolescent rows: **{len(missing_df):,}**"
            )

            if missing_df.empty:
                st.info("No rows found with missing 'School UDISE'. Nothing to export.")
            else:
                # Step 3 (optional): filter by ProgramLaunchName list from second file
                pl_filter_list = []
                if filter_file is not None:
                    with st.spinner("Reading ProgramLaunchName filter file..."):
                        pl_filter_list = get_programlaunch_list_from_file(filter_file)

                    if not pl_filter_list:
                        st.warning(
                            "The ProgramLaunchName filter file did not produce any valid values. "
                            "Proceeding without this filter."
                        )
                    else:
                        st.write(
                            f"ProgramLaunchName values from filter file: **{len(pl_filter_list):,}**"
                        )
                        missing_df = missing_df[
                            missing_df["ProgramLaunchName"].isin(pl_filter_list)
                        ].copy()
                        st.write(
                            f"Rows with missing 'School UDISE' after applying ProgramLaunchName filter: "
                            f"**{len(missing_df):,}**"
                        )

                        if missing_df.empty:
                            st.info(
                                "After applying the ProgramLaunchName filter, "
                                "no rows remain with missing UDISE. Nothing to export."
                            )
                            st.stop()

                # -----------------------------
                # Data-quality / flag columns
                # -----------------------------

                # 1) UDISE_Missing(Yes/No) -> "Missing"/"Available"
                if "School UDISE" in missing_df.columns:
                    udise_series = (
                        missing_df["School UDISE"]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                    )
                    udise_missing_flag = udise_series.eq("")
                    missing_df["UDISE_Missing(Yes/No)"] = udise_missing_flag.map(
                        {True: "Missing", False: "Available"}
                    )

                # 2) Check Child School Name -> blank or "null"
                if "Child School Name" in missing_df.columns:
                    csn = (
                        missing_df["Child School Name"]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                    )
                    child_school_flag = csn.eq("") | csn.str.upper().eq("NULL")
                    missing_df["Check Child School Name"] = child_school_flag.map(
                        {True: "Missing", False: "Available"}
                    )

                # 3) Check Date Of Birth -> flag if DATE OF BIRTH is 1 Jan (1-1- or 01-01-)
                if "DATE OF BIRTH" in missing_df.columns:
                    dob_series = (
                        missing_df["DATE OF BIRTH"]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                    )
                    # Match 1/1 or 01-01 at start of date (before year)
                    dob_flag = dob_series.str.contains(
                        r"\b0?1[-/ ]0?1[-/]", na=False
                    )
                    missing_df["Check Date Of Birth"] = dob_flag.map(
                        {True: "Flag – DOB is 1 Jan", False: ""}
                    )

                # 4) Check Phone Number -> missing or invalid Indian 10-digit mobile (6–9 start)
                if "CONTACTNUMBER" in missing_df.columns:
                    raw_cn = (
                        missing_df["CONTACTNUMBER"]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                    )
                    # Keep only digits for validation
                    digits = raw_cn.str.replace(r"\D", "", regex=True)
                    valid_mobile = digits.str.match(r"^[6-9]\d{9}$", na=False)
                    # Flag if empty or invalid
                    phone_flag = (digits.eq("")) | (~valid_mobile)
                    missing_df["Check Phone Number"] = phone_flag.map(
                        {True: "Missing/Invalid", False: "Valid"}
                    )

                # 5) CASTE interpretation + flag
                if "CASTE" in missing_df.columns:
                    caste_raw = (
                        missing_df["CASTE"]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                    )
                    caste_upper = caste_raw.str.upper()

                    # Interpretation column with meaningful text
                    caste_interp = []
                    for v in caste_upper:
                        if v == "DONT KNOW":
                            caste_interp.append("Caste not known")
                        elif v == "DONT WISH":
                            caste_interp.append("Caste not disclosed")
                        else:
                            caste_interp.append("")
                    missing_df["CASTE_Interpretation"] = caste_interp

                    # Flag column
                    caste_flag = caste_upper.isin(["DONT KNOW", "DONT WISH"])
                    missing_df["Check Caste"] = caste_flag.map(
                        {True: "Caste not known/disclosed", False: ""}
                    )

                # 6) Check Parent Consent -> blank or "null" => Missing/Available
                if "Parent Consent" in missing_df.columns:
                    pc = (
                        missing_df["Parent Consent"]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                    )
                    pc_missing_flag = pc.eq("") | pc.str.upper().eq("NULL")
                    missing_df["Check Parent Consent"] = pc_missing_flag.map(
                        {True: "Missing", False: "Available"}
                    )

                # -----------------------------
                # Summary table: ProgramLaunchName vs count of missing UDISE
                # -----------------------------
                summary_df = (
                    missing_df.groupby("ProgramLaunchName", dropna=False)
                    .size()
                    .reset_index(name="Missing_UDISE_Count")
                    .sort_values("Missing_UDISE_Count", ascending=False)
                )

                st.subheader("Summary: Missing UDISE by ProgramLaunchName")
                st.dataframe(summary_df, use_container_width=True)

                # Prepare output folder
                output_dir = Path(output_folder_input).expanduser()
                try:
                    output_dir.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    st.error(f"Could not create/access output folder: {e}")
                else:
                    groups = missing_df.groupby("ProgramLaunchName", dropna=False)
                    saved_files = []

                    # Description of all checks (for footer)
                    checks_list = [
                        "UDISE code missing",
                        "Child school name missing/null",
                        "Date of birth = 1 Jan",
                        "Phone number missing/invalid (India mobile)",
                        "Caste = DONT KNOW/DONT WISH",
                        "Parent consent missing/null",
                    ]
                    checks_str = ", ".join(checks_list)

                    # Create ZIP in memory
                    zip_buffer = io.BytesIO()

                    with st.spinner("Saving Excel files (combined + ProgramLaunchName-wise)..."):
                        with zipfile.ZipFile(
                            zip_buffer, "w", zipfile.ZIP_DEFLATED
                        ) as zf:

                            # ---------- 1) Combined file: ALL_CPRF_issues ----------
                            if not missing_df.empty:
                                now = datetime.now().strftime("%d-%b-%Y %H:%M")
                                footer_text_all = (
                                    f"Filter: Report created on {now}, "
                                    f"Checks: {checks_str}"
                                )

                                # Create blank row + footer row
                                blank_row_all = {col: "" for col in missing_df.columns}
                                footer_row_all = {col: "" for col in missing_df.columns}
                                footer_row_all[list(missing_df.columns)[0]] = footer_text_all

                                all_with_footer = pd.concat(
                                    [
                                        missing_df,
                                        pd.DataFrame([blank_row_all]),
                                        pd.DataFrame([footer_row_all]),
                                    ],
                                    ignore_index=True,
                                )

                                all_path = output_dir / "ALL_CPRF_issues.xlsx"
                                # Save with styling
                                with pd.ExcelWriter(all_path, engine="openpyxl") as writer:
                                    all_with_footer.to_excel(writer, index=False)
                                    wb = writer.book
                                    ws = writer.sheets[wb.sheetnames[0]]
                                    apply_issue_coloring(ws, missing_df)
                                saved_files.append(str(all_path))

                                # Also add to ZIP
                                sub_buffer_all = io.BytesIO()
                                with pd.ExcelWriter(sub_buffer_all, engine="openpyxl") as writer:
                                    all_with_footer.to_excel(writer, index=False)
                                    wb = writer.book
                                    ws = writer.sheets[wb.sheetnames[0]]
                                    apply_issue_coloring(ws, missing_df)
                                zf.writestr(
                                    "ALL_CPRF_issues.xlsx",
                                    sub_buffer_all.getvalue(),
                                )
                            # -----------------------------------------------------

                            # ---------- 2) Per ProgramLaunchName files ----------
                            for program_name, g in groups:
                                safe_name = safe_filename(program_name)
                                file_path = output_dir / f"{safe_name}.xlsx"

                                now = datetime.now().strftime("%d-%b-%Y %H:%M")
                                footer_text = (
                                    f"Filter: Report created on {now}, "
                                    f"Checks: {checks_str}"
                                )

                                # blank row + footer row for this subset
                                blank_row = {col: "" for col in g.columns}
                                footer_row = {col: "" for col in g.columns}
                                footer_row[g.columns[0]] = footer_text

                                g_with_footer = pd.concat(
                                    [g, pd.DataFrame([blank_row]), pd.DataFrame([footer_row])],
                                    ignore_index=True,
                                )

                                # Save to disk with styling
                                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                                    g_with_footer.to_excel(writer, index=False)
                                    wb = writer.book
                                    ws = writer.sheets[wb.sheetnames[0]]
                                    apply_issue_coloring(ws, g)
                                saved_files.append(str(file_path))

                                # Also add to ZIP (in-memory) with styling
                                sub_buffer = io.BytesIO()
                                with pd.ExcelWriter(sub_buffer, engine="openpyxl") as writer:
                                    g_with_footer.to_excel(writer, index=False)
                                    wb = writer.book
                                    ws = writer.sheets[wb.sheetnames[0]]
                                    apply_issue_coloring(ws, g)
                                zf.writestr(f"{safe_name}.xlsx", sub_buffer.getvalue())
                            # -----------------------------------------------------

                    # ✅ Increment run counter only after successful export
                    new_count = increment_run_counter()
                    st.sidebar.metric("Total runs (all time)", new_count)
                    st.info(f"This tool has been run {new_count} time(s) in total.")

                    st.success(
                        f"Exported {len(saved_files)} Excel file(s) to:\n\n`{output_dir}`"
                    )

                    st.write("Sample of output paths:")
                    for fp in saved_files[:10]:
                        st.write("- ", fp)
                    if len(saved_files) > 10:
                        st.write(f"... and {len(saved_files) - 10} more files.")

                    # Prepare ZIP for download
                    zip_buffer.seek(0)
                    st.subheader("Download all files as ZIP")
                    st.download_button(
                        label="Download ZIP (ALL_CPRF_issues + ProgramLaunchName files)",
                        data=zip_buffer.getvalue(),
                        file_name="cprf_dq_by_programlaunch.zip",
                        mime="application/zip",
                    )